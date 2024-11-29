# Importación de las bibliotecas necesarias
import time
from selenium import webdriver
from bs4 import BeautifulSoup
import undetected_chromedriver as uc
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import numpy as np
from sklearn.preprocessing import MinMaxScaler
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Listas para almacenar datos
hrefs = []  # Enlaces a las páginas de los equipos
playerList = []  # Información detallada de los jugadores

# Configuración de las opciones del navegador
options = Options()
options.add_argument("start-maximized")  # Maximizar ventana al iniciar
options.add_argument("disable-infobars")  # Deshabilitar mensajes informativos del navegador
options.add_argument("--disable-extensions")  # Deshabilitar extensiones
options.add_argument("--disable-gpu")  # Deshabilitar aceleración por GPU
options.add_argument("--no-sandbox")  # Deshabilitar el sandbox
# options.add_argument("--headless")  # Ejecutar navegador en modo headless (sin interfaz gráfica)
options.add_argument("--disable-crash-reporter")  # Deshabilitar reportes de fallos
options.add_argument("--disable-popup-blocking")  # Deshabilitar bloqueo de ventanas emergentes

# Inicializar navegador con opciones configuradas
driver = uc.Chrome(options=options)

# Abrir página principal de estadísticas de equipos en HLTV
driver.get('https://www.hltv.org/stats/teams?startDate=2024-01-01&endDate=2024-12-31&rankingFilter=Top50')

# Guardar el identificador de la ventana actual
windows_teamList = driver.current_window_handle

# Esperar a que los elementos del selector de equipos sean visibles
elements = WebDriverWait(driver, 30).until(
    EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "td.teamCol-teams-overview>a"))
)
time.sleep(3)  # Pausa adicional para garantizar que los elementos estén cargados

# Función para extraer valores de estadísticas mediante XPATH
def unpack(xpath):
    stat_value = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, xpath))
    ).text
    return stat_value

# Recopilar enlaces de los equipos
for element in elements:
    hrefs.append(element.get_attribute("href"))

# Recorrer cada equipo para obtener detalles
for href in hrefs: # Añade [:número] para eligir hasta el equipo que quieres que scrapee
    driver.execute_script("window.open('" + href + "');")  # Abrir enlace en una nueva pestaña
    windows_after = driver.window_handles
    window_Team = [x for x in windows_after if x != windows_teamList][1]  # Identificar nueva pestaña
    driver.switch_to.window(window_Team)  # Cambiar a la pestaña del equipo

    # Hacer clic en la pestaña "Lineups"
    lineups = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), 'Lineups')]"))
    ).click()

    # Scraping de datos de alineaciones
    bs = BeautifulSoup(driver.page_source, "lxml")
    lineupContainers = bs.find_all("div", class_="lineup-container")  # Contenedores de alineaciones
    teamName = (bs.find("span", class_="context-item-name")).text  # Nombre del equipo

    for lineupContainer in lineupContainers:
        lineupYear = ((lineupContainer.find("div", class_="lineup-year")).text).replace('Replace context with lineup', '')
        mapsPlusWinrate = lineupContainer.find_all("div", class_="large-strong")
        mapsPlayed = int(mapsPlusWinrate[0].text)  # Mapas jugados
        winrate = round(int(mapsPlusWinrate[1].text.split()[0]) / mapsPlayed, 2)  # Winrate

        # Scraping de jugadores en la alineación
        lineupMembers = WebDriverWait(driver, 30).until(
            EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "div.teammate-info>a"))
        )
        newhrefs = [member.get_attribute("href") for member in lineupMembers]
        playerPrelimNames = [member.find_element(By.CSS_SELECTOR, "div.text-ellipsis").text for member in lineupMembers]

        # Recopilar estadísticas de cada jugador
        for j, newhref in enumerate(newhrefs):
            playerPrelimName = playerPrelimNames[j]

            # Procesar solo si el jugador aún no está en la lista
            if not any(d['playerName'] == playerPrelimName for d in playerList):
                valueToFill = True
                driver.execute_script("window.open('" + newhref + "');")  # Abrir página del jugador
                windows_after = driver.window_handles
                window_Player = [x for x in windows_after if x != window_Team][2]  # Identificar pestaña del jugador
                driver.switch_to.window(window_Player)  # Cambiar a la pestaña del jugador

                # Scraping de estadísticas del jugador
                pbs = BeautifulSoup(driver.page_source, "lxml")
                playerName = pbs.find("h1", class_="summaryNickname").text
                stats = pbs.find_all("div", class_="stats-row")
                statTopPanel = pbs.find_all("div", class_="summaryStatBreakdownDataValue")
                statImpact = statTopPanel[3].text
                statKAST = (statTopPanel[2].text).strip("%")

                statList = {}
                for stat in stats:
                    statPair = stat.find_all("span")
                    actualStat = (statPair[1].text).strip("%") if valueToFill else 'NA'
                    statList = {**statList, **dict({f"{statPair[0].text}": actualStat})}

                if valueToFill:
                    # Obtener estadísticas adicionales individuales
                    WebDriverWait(driver, 30).until(
                        EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), 'Individual')]"))
                    ).click()

                    statOKratio = unpack('/html/body/div[2]/div[8]/div[2]/div[1]/div/div[6]/div/div[1]/div[5]/div[3]/span[2]')
                    statOK = unpack('/html/body/div[2]/div[8]/div[2]/div[1]/div/div[6]/div/div[1]/div[5]/div[1]/span[2]')
                    statRifleKills = unpack("/html/body/div[2]/div[8]/div[2]/div[1]/div/div[6]/div/div[2]/div[5]/div[1]/span[2]")
                    statSniperKills = unpack("/html/body/div[2]/div[8]/div[2]/div[1]/div/div[6]/div/div[2]/div[5]/div[2]/span[2]")
                    statSmgKills = unpack("/html/body/div[2]/div[8]/div[2]/div[1]/div/div[6]/div/div[2]/div[5]/div[3]/span[2]")
                    statPistolKills = unpack("/html/body/div[2]/div[8]/div[2]/div[1]/div/div[6]/div/div[2]/div[5]/div[4]/span[2]")

                # Añadir estadísticas a la lista
                playerList.append({
                    "playerName": playerName,
                    "teamName": f"{teamName} {lineupYear}",
                    "mapsPlayed": mapsPlayed,
                    "winrate": winrate,
                    "hltvImpact": statImpact,
                    "KAST": statKAST,
                    "openingKills": statOK,
                    "OKratio": statOKratio,
                    "rifleKills": statRifleKills,
                    "sniperKills": statSniperKills,
                    "smgKills": statSmgKills,
                    "pistolKills": statPistolKills,
                    **statList
                })

                driver.close()  # Cerrar pestaña del jugador
            driver.switch_to.window(window_Team)  # Volver a la pestaña del equipo
    driver.close()  # Cerrar pestaña del equipo
    driver.switch_to.window(windows_teamList)  # Volver a la pestaña principal

driver.quit()  # Cerrar el navegador


# Guardar los datos en un archivo xlsx
df = pd.DataFrame(playerList)
df.replace('NA', np.nan, regex=True, inplace=True)  # Reemplazar 'NA' con valores nulos
# Definir las columnas que se deben excluir de la limpieza
columns_to_exclude = ['playerName', 'teamName']
# Paso 1: Identificar columnas numéricas, excluyendo las especificadas
columns_to_clean = [col for col in df.columns if col not in columns_to_exclude]
# Paso 2: Convertir a numérico las columnas seleccionadas
for column in columns_to_clean:
    df[column] = pd.to_numeric(df[column], errors='coerce')
# Verificar valores faltantes después de la limpieza
missing_values_after_cleaning = df.isna().sum()


# Generar nuevas columnas y crear nuevas estadísticas con los datos obtenidos
df['KILLS_PER_MAP'] = df['Total kills'] / df['mapsPlayed'] # Kills por mapa
df['DEATHS_PER_MAP'] = df['Total deaths'] / df['mapsPlayed'] # Muertes por mapa
df['KILLS_TO_ROUND_RATIO'] = df['Kills / round'] * df['Rounds played'] # Relación kills/round
df['OPENING_EFFICIENCY'] = df['openingKills'] / df['Total kills'] # Apertura efectiva
df['RIFLE_KILL_RATIO'] = df['rifleKills'] / df['Total kills'] # Proporción de kills para rifler
df['SNIPER_KILL_RATIO'] = df['sniperKills'] / df['Total kills'] # Proporción de kills para sniper
df['SMG_KILL_RATIO'] = df['smgKills'] / df['Total kills'] # Proporción de kills para smg
df['PISTOL_KILL_RATIO'] = df['pistolKills'] / df['Total kills'] # Proporción de kills para pistol
df['SAVED_EFFICIENCY'] = (df['Saved by teammate / round'] + df['Saved teammates / round']) # Eficiencia de salvamento
df['GRENADE_IMPACT'] = df['Grenade dmg / Round'] / df['Kills / round'] # Impacto de granadas por kill
df['TEAM_CONTRIBUTION_PER_ROUND'] = (df['Kills / round'] + df['Assists / round'] - df['Deaths / round']) # Aportes al equipo por ronda (kills + assists - deaths)


# GENERAMOS NUESTRA F1_SCORE
key_columns = ['winrate', 'K/D Ratio', 'Rating 2.1', 'hltvImpact'] # Selección de columnas clave
# Crear un escalador y ajustar-transformar los datos
scaler = MinMaxScaler()
normalized_data = scaler.fit_transform(df[key_columns])
normalized_df = pd.DataFrame(normalized_data, columns=key_columns) # Crear un DataFrame con los datos normalizados
# Calcular un puntaje de desempeño único como un promedio ponderado (puedes ajustar los pesos según prioridad)
weights = [0.3, 0.3, 0.2, 0.2]  # Pesos para las columnas: winrate, K/D Ratio, Rating 2.1, hltvImpact
normalized_df['SCORE_PERFORMANCE'] = (normalized_df * weights).sum(axis=1)
df['SCORE_PERFORMANCE'] = normalized_df['SCORE_PERFORMANCE'] # Agregar el puntaje único al dataset original


# Renombrar columnas del dataframe
df.rename(columns={'playerName': 'PLAYER',
                   'teamName': 'TEAM',
                   'mapsPlayed':'MAPS_PLAYED',
                   'winrate': 'WINRATE',
                   'hltvImpact': 'HLTV-IMPACT',
                   'KAST': 'KAST',
                   'openingKills': 'OPENING_Kills',
                   'OKratio': 'OK_Ratio',
                   'rifleKills': 'RIFLE_Kills',
                   'sniperKills': 'SNIPER_Kills',
                   'smgKills': 'SMG_Kills',
                   'pistolKills': 'PISTOL_Kills',
                   'Total kills': 'TOTAL_Kills',
                   'Headshot %': '%HEADSHOT',
                   'Total deaths': 'TOTAL_Deaths',
                   'K/D Ratio': 'K/D_Ratio',
                   'Damage / Round': 'DAMAGE/Round',
                   'Grenade dmg / Round': 'GRANADE_DMG/Round',
                   'Maps played': 'MAPS_PLAYED',
                   'Rounds played': 'ROUNDS_PLAYED',
                   'Kills / round': 'KILLS/Round',
                   'Assists / round': 'ASSISTS/Round',
                   'Deaths / round': 'DEATHS/Rounds',
                   'Saved by teammate / round': 'SAVED BY TEAMMATE/Round',
                   'Saved teammates / round': 'SAVED TEAMMATES/Round',
                   'Rating 2.1': 'RATING 2.1'}, inplace=True)
df.sort_values(by='PLAYER', inplace=True)  # Ordenar por nombre del jugador
#df.to_excel("hltv-stats-before-major-2024_1.xlsx", index=False)  # Guardado del archivo en formato .xlsx


# Guardar el DataFrame en un archivo de Excel
output_path = 'hltv-stats-before-major-2024.xlsx'
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='STATS')
    # Acceder al libro y la hoja
    workbook = writer.book
    worksheet = workbook['STATS']
    # Definir los colores de fondo
    color_1 = PatternFill(start_color="364250", end_color="364250", fill_type="solid")
    color_2 = PatternFill(start_color="2D3844", end_color="2D3844", fill_type="solid")
    font_color = Font(color="929A9E") # Definir el color de la fuente
    bold_font = Font(bold=True, color="929A9E")  # Fuente en Negrita y color gris para los encabezados
    worksheet.auto_filter.ref = worksheet.dimensions # Aplicar autofiltro en todas las columnas
    # Ajuste automático de las columnas (AutoFit)
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtenemos la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 5.5) # Ajuste del tamaño de cada celda
        worksheet.column_dimensions[column].width = adjusted_width
    # Alternar colores de fondo y cambiar color de la fuente
    for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column),
            start=2):
        # Alternar colores de fondo
        fill_color = color_1 if row_index % 2 == 0 else color_2
        for cell in row:
            cell.fill = fill_color
            cell.font = font_color  # Cambiar color de la fuente a gris
    # Aplicar el estilo a la primera fila (negrita y color de fuente)
    for cell in worksheet[1]:  # Primera fila: Encabezados en Negrita
        cell.font = bold_font

print(f"HLTV.org - Scraper is DONE: {output_path}")
